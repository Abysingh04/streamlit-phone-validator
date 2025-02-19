import requests
import pandas as pd
import time
import os
from openpyxl import load_workbook

# File paths
input_file = "D:/Python/Streamlit/Framework/InputPhoneNumber.xlsx"
output_file = "output.xlsx"

# Check if input file exists
if not os.path.exists(input_file):
    print(f"❌ Error: Input file '{input_file}' not found!")
    exit()

# Read Excel file as text (to prevent formatting issues)
try:
    df = pd.read_excel(input_file, dtype=str, sheet_name="Sheet2")
    df.columns = df.columns.str.strip()  # Remove extra spaces
except Exception as e:
    print(f"❌ Error loading Excel file: {e}")
    exit()

# Ensure "Phone_Number" column exists
if "Phone_Number" not in df.columns:
    print(f"❌ Error: 'Phone_Number' column is missing! Available columns: {df.columns}")
    exit()

print("🚀 Starting phone number validation process...")

# API details
api_base_url = "http://phone-number-api.com/csv/"
fields = "status,message,numberType,numberValid,numberValidForRegion,isDisposible,numberCountryCode,numberAreaCode,formatE164,formatInternational,carrier,continent,countryName,country,region,regionName,city,zip,query"

# Expected Response Headers
expected_headers = [
    "Status", "Message", "Number Type", "Number Valid","numberValidForRegion", "Is Disposable",
    "Country Code", "Area Code", "E164 Format", "International Format",
    "Carrier", "Continent", "Country Name", "Country", "Region",
    "Region Name", "City", "ZIP", "Query"
]

# Function to append data incrementally
def append_to_excel(data, file_path, sheet_name="Sheet1"):
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            workbook = load_workbook(file_path)
            sheet = writer.sheets[sheet_name]
            start_row = sheet.max_row  # Find next empty row
            data.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        data.to_excel(file_path, index=False)  # If file doesn't exist, create new

# Results List
results = []
batch_size = 5  # Save after every 5 requests

# Loop through each phone number
for index, phone_number in enumerate(df["Phone_Number"], start=1):
    if pd.isna(phone_number) or phone_number.strip() == "":
        print("⚠️ Skipping empty phone number.")
        continue

    phone_number = phone_number.strip()

    # Ensure the phone number starts with "+"
    if not phone_number.startswith("+"):
        print(f"⚠️ Warning: '{phone_number}' is missing '+'. Skipping...")
        continue

    # Construct API Request URL
    url = f"{api_base_url}?number={phone_number}&fields={fields}"
    print(f"📡 Sending request {index}: {phone_number} → {url}")

    try:
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            values = response.text.strip().split(",")

            # Handle Column Mismatches
            if len(values) < len(expected_headers):
                values += ["N/A"] * (len(expected_headers) - len(values))
            elif len(values) > len(expected_headers):
                print(f"⚠️ Extra fields in API response for {phone_number}. Adjusting...")
                expected_headers = expected_headers[:len(values)]

            # Store Result
            results.append(values)
        else:
            print(f"❌ API Error {response.status_code} for {phone_number}: {response.text}")

        # ⏳ Maintain 5 Requests per Minute (12-sec delay)
        time.sleep(12)

        # Save every `batch_size` requests
        if index % batch_size == 0 and results:
            print(f"💾 Saving {len(results)} new results to '{output_file}'...")
            new_data = pd.DataFrame(results, columns=expected_headers)
            append_to_excel(new_data, output_file)
            results = []  # Clear results list

    except requests.exceptions.RequestException as e:
        print(f"⚠️ Request failed for {phone_number}: {e}")

# Final Save if any remaining results
if results:
    print(f"💾 Final saving {len(results)} remaining results...")
    new_data = pd.DataFrame(results, columns=expected_headers)
    append_to_excel(new_data, output_file)

print("\n✅ Process complete! Data saved incrementally.")
