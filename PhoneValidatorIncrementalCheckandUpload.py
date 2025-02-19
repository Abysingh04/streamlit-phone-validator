import requests
import pandas as pd
import time
import os
from openpyxl import load_workbook

# 📂 File Paths
input_file = "D:/Python/Streamlit/Framework/PhoneValidator/InputPhoneNumber.xlsx"
output_file = "D:/Python/Streamlit/Framework/PhoneValidator/output_Incremental.xlsx"

# 🛑 Check if input file exists
if not os.path.exists(input_file):
    print(f"❌ Error: Input file '{input_file}' not found!")
    exit()

# 📥 Read Input File
try:
    df = pd.read_excel(input_file, dtype=str, sheet_name="Sheet2")
    df.columns = df.columns.str.strip()  # Remove extra spaces in column names
except Exception as e:
    print(f"❌ Error loading Excel file: {e}")
    exit()

# Ensure "Phone_Number" column exists
if "Phone_Number" not in df.columns:
    print(f"❌ Error: 'Phone_Number' column is missing! Available columns: {df.columns}")
    exit()

print("🚀 Starting phone number validation process...")

# 📝 Load Already Processed Phone Numbers
processed_numbers = set()
if os.path.exists(output_file):
    try:
        df_output = pd.read_excel(output_file, dtype=str)
        if df_output.empty:
            print(f"⚠️ Warning: '{output_file}' exists but is empty.")
        else:
            print(f"✅ Output File Loaded Successfully!")

        # Normalize stored numbers by removing spaces and dashes
        if "Query" in df_output.columns:
            processed_numbers = set(df_output["Query"].dropna().str.strip().str.replace(" ", "").str.replace("-", ""))
            print(f"🔍 Found {len(processed_numbers)} previously processed phone numbers. Skipping them...")
        else:
            print(f"❌ Error: 'Query' column not found in '{output_file}'. Available columns: {df_output.columns}")

    except Exception as e:
        print(f"⚠️ Warning: Could not read '{output_file}': {e}")
else:
    print(f"⚠️ Warning: Output file '{output_file}' does not exist yet.")

# 🌐 API Details
api_base_url = "http://phone-number-api.com/csv/"
fields = "status,numberType,numberValid,numberValidForRegion,isDisposible,numberCountryCode,numberAreaCode,formatE164,formatNational,formatInternational,carrier,continent,continentCode,countryName,country,region,regionName,city,zip,offset,currency,query"

# Expected API Response Headers
expected_headers = [
    "Status",  "Number Type", "Number Valid", "numberValidForRegion", "Is Disposable",
    "Country Code", "Area Code", "E164 Format", "National Format", "International Format",
    "Carrier", "Continent", "Continent Code", "Country Name", "Country", "Region",
    "Region Name", "City", "ZIP", "Offset", "Currency", "Query"
]

# 📌 Function to append data to Excel
def append_to_excel(data, file_path, sheet_name="Sheet1"):
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            workbook = load_workbook(file_path)
            sheet = writer.sheets.get(sheet_name, writer.book.active)
            start_row = sheet.max_row  # Find next empty row
            data.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        data.to_excel(file_path, index=False)  # Create new file if it doesn't exist

# 🏁 Results Storage
results = []
batch_size = 5  # Save after every 5 requests

# 🔄 Processing Phone Numbers
for index, phone_number in enumerate(df["Phone_Number"], start=1):
    if pd.isna(phone_number) or phone_number.strip() == "":
        print("⚠️ Skipping empty phone number.")
        continue

    phone_number = phone_number.strip().replace(" ", "").replace("-", "")  # Normalize format

    # Check if already processed
    if phone_number in processed_numbers:
        print(f"✅ Skipping already processed: {phone_number}")
        continue

    # Ensure phone number starts with "+"
    if not phone_number.startswith("+"):
        print(f"⚠️ Warning: '{phone_number}' is missing '+'. Skipping...")
        continue

    # 📡 Construct API Request
    url = f"{api_base_url}?number={phone_number}&fields={fields}"
    print(f"📡 Sending request {index}: {phone_number}")

    try:
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            values = response.text.strip().split(",")

            # Ensure the response has exactly 23 fields (fill missing with "N/A")
            while len(values) < len(expected_headers):
                values.append("N/A")

            if len(values) > len(expected_headers):
                values = values[:len(expected_headers)]  # Trim extra fields

        else:
            print(f"❌ API Error: {response.status_code} for {phone_number}")
            values = ["API_ERROR"] + ["N/A"] * (len(expected_headers) - 2) + [phone_number]  # Fill error & copy phone number

    except requests.exceptions.RequestException as e:
        print(f"⚠️ Request failed for {phone_number}: {e}")
        values = ["REQUEST_FAILED"] + ["N/A"] * (len(expected_headers) - 2) + [phone_number]  # Copy phone number

    # ✅ Ensure "Query" column is always filled with the original phone number
    if values[-1] == "N/A":
        values[-1] = phone_number

    # ✅ Store processed number
    results.append(values)

    # ⏳ Rate Limit (5 Requests per Minute → 12-sec delay)
    time.sleep(12)

    # 💾 Save every `batch_size` requests
    if index % batch_size == 0 and results:
        print(f"💾 Saving {len(results)} new results to '{output_file}'...")
        new_data = pd.DataFrame(results, columns=expected_headers)
        append_to_excel(new_data, output_file)
        results = []  # Clear results list

# 🏁 Final Save for remaining results
if results:
    print(f"💾 Final saving {len(results)} remaining results...")
    new_data = pd.DataFrame(results, columns=expected_headers)
    append_to_excel(new_data, output_file)

print("\n✅ Process complete! Data saved incrementally.")
