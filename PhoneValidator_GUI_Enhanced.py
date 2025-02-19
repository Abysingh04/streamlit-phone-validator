import streamlit as st
import pandas as pd
import time
import requests
import io
import os
from openpyxl import load_workbook

# 🌐 API Details
API_BASE_URL = "http://phone-number-api.com/csv/"
FIELDS = "status,numberType,numberValid,numberValidForRegion,isDisposible,numberCountryCode,numberAreaCode,formatE164,formatNational,formatInternational,carrier,continent,continentCode,countryName,country,region,regionName,city,zip,offset,currency,query"

# Expected API Response Headers
EXPECTED_HEADERS = [
    "Status", "Number Type", "Number Valid", "numberValidForRegion", "Is Disposable",
    "Country Code", "Area Code", "E164 Format", "National Format", "International Format",
    "Carrier", "Continent", "Continent Code", "Country Name", "Country", "Region",
    "Region Name", "City", "ZIP", "Offset", "Currency", "Query"
]

OUTPUT_FILE = "validated_numbers.xlsx"

# 📌 Function to append data to Excel
def append_to_excel(data, file_path, sheet_name="Sheet1"):
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            workbook = load_workbook(file_path)
            sheet = writer.sheets.get(sheet_name, writer.book.active)
            start_row = sheet.max_row  # Find next empty row
            data.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        data.to_excel(file_path, index=False)  # Create new file if it doesn't exist

# Streamlit UI
st.title("📞 Phone Number Validator")

uploaded_file = st.file_uploader("Upload an Excel file with phone numbers", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file, dtype=str)
    df.columns = df.columns.str.strip()
    
    if "Phone_Number" not in df.columns:
        st.error("❌ 'Phone_Number' column is missing! Please upload a valid file.")
    else:
        st.success("✅ File uploaded successfully!")
        
        # 📝 Load Already Processed Phone Numbers
        processed_numbers = set()
        if os.path.exists(OUTPUT_FILE):
            try:
                df_output = pd.read_excel(OUTPUT_FILE, dtype=str)
                if "Query" in df_output.columns:
                    processed_numbers = set(df_output["Query"].dropna().str.strip().str.replace(" ", "").str.replace("-", ""))
            except:
                pass
        
        results = []
        progress_bar = st.progress(0)
        total_numbers = len(df)
        
        for index, phone_number in enumerate(df["Phone_Number"].dropna().astype(str), start=1):
            phone_number = phone_number.strip().replace(" ", "").replace("-", "")
            
            if phone_number in processed_numbers:
                continue
            
            if not phone_number.startswith("+"):
                results.append(["INVALID_FORMAT"] + ["N/A"] * (len(EXPECTED_HEADERS) - 2) + [phone_number])
                continue
            
            url = f"{API_BASE_URL}?number={phone_number}&fields={FIELDS}"
            
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    values = response.text.strip().split(",")
                    values += ["N/A"] * (len(EXPECTED_HEADERS) - len(values))
                else:
                    values = ["API_ERROR"] + ["N/A"] * (len(EXPECTED_HEADERS) - 2) + [phone_number]
            except:
                values = ["REQUEST_FAILED"] + ["N/A"] * (len(EXPECTED_HEADERS) - 2) + [phone_number]
                
            results.append(values)
            progress_bar.progress(index / total_numbers)
            time.sleep(12)  # Enforcing API rate limit
            
            if index % 5 == 0 and results:
                new_data = pd.DataFrame(results, columns=EXPECTED_HEADERS)
                append_to_excel(new_data, OUTPUT_FILE)
                results = []  # Clear results list
        
        if results:
            new_data = pd.DataFrame(results, columns=EXPECTED_HEADERS)
            append_to_excel(new_data, OUTPUT_FILE)
        
        st.success("✅ Processing complete! Data saved incrementally.")
        
        with open(OUTPUT_FILE, "rb") as f:
            st.download_button(
                label="📥 Download Processed File",
                data=f,
                file_name=OUTPUT_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def main():
    st.write("**Important Instruction**: Your file must have a column called **< Phone_Number >** to work this code")

if __name__ == "__main__":
    main()
